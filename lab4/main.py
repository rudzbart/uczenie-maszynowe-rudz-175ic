from openpyxl import Workbook
from bs4 import BeautifulSoup
import requests
import string
import random

wb = Workbook()

ws1 = wb.create_sheet("Giełda")
ws2 = wb.create_sheet("Linki")
ws3 = wb.create_sheet("Filmweb")


def random_char(y):
    return ''.join(random.choice(string.ascii_lowercase) for x in range(y))


firma = []
kurs = []
zmiana = []
# transakcje = ["Test"]
i = 0
while i < 5:
    randomLetters = random_char(3)

    page = requests.get("https://stooq.pl/q/?s=" + randomLetters)
    soup = BeautifulSoup(page.content, 'html.parser')
    if soup.title.string != "Wyszukiwanie symbolu - Stooq":
        print(soup.title.string)
        if soup.find(id="aq_" + randomLetters + "_c0") is None and soup.find(
                id="aq_" + randomLetters + "_c1") is None and soup.find(
            id="aq_" + randomLetters + "_c2") is None and soup.find(id="aq_" + randomLetters + "_c3") is None:
            print("Kurs: " + soup.find(id="aq_" + randomLetters + "_c4").get_text())
            kurs.append(soup.find(id="aq_" + randomLetters + "_c4").get_text())
        elif soup.find(id="aq_" + randomLetters + "_c1") is None and soup.find(
                id="aq_" + randomLetters + "_c2") is None and soup.find(
            id="aq_" + randomLetters + "_c4") is None and soup.find(id="aq_" + randomLetters + "_c0") is None:
            print("Kurs: " + soup.find(id="aq_" + randomLetters + "_c3").get_text())
            kurs.append(soup.find(id="aq_" + randomLetters + "_c3").get_text())
        elif soup.find(id="aq_" + randomLetters + "_c1") is None and soup.find(
                id="aq_" + randomLetters + "_c3") is None and soup.find(
            id="aq_" + randomLetters + "_c4") is None and soup.find(id="aq_" + randomLetters + "_c0") is None:
            print("Kurs: " + soup.find(id="aq_" + randomLetters + "_c2").get_text())
            kurs.append(soup.find(id="aq_" + randomLetters + "_c2").get_text())
        elif soup.find(id="aq_" + randomLetters + "_c2") is None and soup.find(
                id="aq_" + randomLetters + "_c3") is None and soup.find(
            id="aq_" + randomLetters + "_c4") is None and soup.find(id="aq_" + randomLetters + "_c0") is None:
            print("Kurs: " + soup.find(id="aq_" + randomLetters + "_c1").get_text())
            kurs.append(soup.find(id="aq_" + randomLetters + "_c1").get_text())
        elif soup.find(id="aq_" + randomLetters + "_c1") is None and soup.find(
                id="aq_" + randomLetters + "_c2") is None and soup.find(
            id="aq_" + randomLetters + "_c3") is None and soup.find(id="aq_" + randomLetters + "_c4") is None:
            print("Kurs: " + soup.find(id="aq_" + randomLetters + "_c0").get_text())
            kurs.append(soup.find(id="aq_" + randomLetters + "_c0").get_text())
        print("Zmiana w %: " + soup.find(id="aq_" + randomLetters + "_m3").get_text())
        zmiana.append(soup.find(id="aq_" + randomLetters + "_m3").get_text())
        firma.append(soup.title.string)
        # print(soup.find(id="aq_" + randomLetters + "_n1"))
        # print("Liczba transakcji: " + soup.find(id="aq_" + randomLetters + "_n1").get_text())
        # transakcje[i] = soup.find(id="aq_" + randomLetters + "_n1").get_text()
        i += 1

    else:
        print("Wygenerowane 3 litery nie dały poprawnego rezultatu.")

a1 = ws1['A1']
b1 = ws1['B1']
c1 = ws1['C1']
d1 = ws1['D1']

a2 = ws1['A2']
b2 = ws1['B2']
c2 = ws1['C2']
d2 = ws1['D2']

a1.value = "Firma"
b1.value = "Kurs"
c1.value = "Zmiana w %"
d1.value = "Transakcje"

i = 2
j = 0
while i < 7:
    ws1.cell(row=i, column=1).value = firma[j]
    j += 1
    i += 1

i = 2
j = 0
while i < 7:
    ws1.cell(row=i, column=2).value = kurs[j]
    j += 1
    i += 1

i = 2
j = 0
while i < 7:
    ws1.cell(row=i, column=3).value = zmiana[j]
    j += 1
    i += 1

# i = 2
# j = 0
# while i < 7:
#     # ws1.cell(row=i, column=4).value = transakcje[j]
#     j += 1
#     i += 1

page = requests.get("https://www.gry-online.pl/gry/")
soup = BeautifulSoup(page.content, 'html.parser')

i = 1
for a in soup.find_all('a', href=True):
    ws2.cell(row=i, column=1).value = a['href']
    i += 1

page = requests.get("https://www.filmweb.pl/film/Braveheart+-+Waleczne+Serce-1995-1052")
soup = BeautifulSoup(page.content, 'html.parser')
director = soup.find(text="reżyseria").next_element.find("a").text.strip()
date = soup.find(text="premiera").next_element.find("a").text.strip()
boxOffice = soup.find(text="boxoffice").next_element.text
rating = soup.find("span", itemprop="ratingValue").text.strip()

a1 = ws3['A1']
b1 = ws3['B1']
c1 = ws3['C1']
d1 = ws3['D1']

a2 = ws3['A2']
b2 = ws3['B2']
c2 = ws3['C2']
d2 = ws3['D2']

a1.value = "Reżyser"
b1.value = "Data premiery"
c1.value = "Box Office"
d1.value = "Ocena"

a2.value = director
b2.value = date
c2.value = boxOffice
d2.value = rating

wb.save('Rudz-175IC.xlsx')